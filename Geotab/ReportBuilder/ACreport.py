import mygeotab
import getpass
import requests
import pandas as pd
import datetime
import os
from flask import Flask, jsonify, send_file, request
from functools import lru_cache
from threading import Thread
from dateutil import parser
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import traceback

app = Flask(__name__)

GEO_USER = "username@email.com"
GEO_PASSWORD = getpass.getpass("Enter your Geotab password: ")
GEO_DATABASE = "database"  # Replace with your Geotab database name

geotab_api = None

AVAILABLE_COLUMNS = {
    "Rule": [
        {"id": "activeFrom", "name": "Start Time", "source": "exception"},
        {"id": "activeTo", "name": "End Time", "source": "exception"},
        {"id": "distance", "name": "Distance", "source": "exception"},
        {"id": "duration", "name": "Duration", "source": "exception"},
        {"id": "rule.id", "name": "Rule Name", "source": "exception"},
        {"id": "state", "name": "State", "source": "exception"},
        {"id": "version", "name": "Version", "source": "exception"},
        {"id": "lastModifiedDateTime", "name": "Last Modified", "source": "exception"},
        {"id": "diagnostic", "name": "Diagnostic", "source": "exception"}
    ],
    "Asset": [
        {"id": "name", "name": "Asset Name", "source": "device"},
        {"id": "vehicleIdentificationNumber", "name": "VIN", "source": "device"},
        {"id": "make", "name": "Make", "source": "device"},
        {"id": "model", "name": "Model", "source": "device"},
        {"id": "year", "name": "Year", "source": "device"},
        {"id": "manufacturer", "name": "Manufacturer", "source": "device"},
        {"id": "plant", "name": "Manufacturing Plant", "source": "device"},
        {"id": "vehicleType", "name": "Vehicle Type", "source": "device"},
        {"id": "bodyStyle", "name": "Body Style", "source": "device"},
        {"id": "engineSize", "name": "Engine Size (L)", "source": "device"},
        {"id": "engineCylinders", "name": "Engine Cylinders", "source": "device"},
        {"id": "fuelType", "name": "Fuel Type", "source": "device"},
        {"id": "series", "name": "Series", "source": "device"},
        {"id": "transmission", "name": "Transmission", "source": "device"},
        {"id": "driveType", "name": "Drive Type", "source": "device"},
        {"id": "gvwr", "name": "GVWR", "source": "device"},
        {"id": "engineVehicleIdentificationNumber", "name": "Engine VIN", "source": "device"},
        {"id": "licensePlate", "name": "License Plate", "source": "device"},
        {"id": "licenseState", "name": "License State", "source": "device"},
        {"id": "comment", "name": "Device Comment", "source": "device"},
        {"id": "serialNumber", "name": "Serial Number", "source": "device"},
        {"id": "deviceType", "name": "Device Type", "source": "device"},
        {"id": "workTime", "name": "Work Time", "source": "device"},
        {"id": "timeZoneId", "name": "Time Zone", "source": "device"},
        {"id": "productId", "name": "Product ID", "source": "device"},
        {"id": "groups", "name": "Device Groups", "source": "device"}
    ],
    "Driver": [
        {"id": "firstName", "name": "First Name", "source": "user"},
        {"id": "lastName", "name": "Last Name", "source": "user"},
        {"id": "name", "name": "Username", "source": "user"},
        {"id": "employeeNo", "name": "Employee Number", "source": "user"},
        {"id": "licenseNumber", "name": "License Number", "source": "user"},
        {"id": "licenseProvince", "name": "License Province", "source": "user"},
        {"id": "phoneNumber", "name": "Phone Number", "source": "user"},
        {"id": "companyName", "name": "Company Name", "source": "user"},
        {"id": "companyAddress", "name": "Company Address", "source": "user"},
        {"id": "designation", "name": "Designation", "source": "user"},
        {"id": "comment", "name": "User Comment", "source": "user"},
        {"id": "driverGroups", "name": "Driver Groups", "source": "user"},
        {"id": "timeZoneId", "name": "Time Zone", "source": "user"}
    ]
}

FAULTS_COLUMNS = {
    "Fault": [
        {"id": "dateTime", "name": "Date Time", "source": "faultdata"},
        {"id": "faultState", "name": "Current Status", "source": "faultdata"},
        {"id": "name", "name": "Fault Name", "source": "diagnostic"},
        {"id": "faultCode", "name": "Fault Code", "source": "diagnostic"}
    ],
    "Asset": AVAILABLE_COLUMNS["Asset"],  # Reuse existing device columns
    "Driver": AVAILABLE_COLUMNS["Driver"]  # Reuse existing user columns
}

# Add this after your existing AVAILABLE_COLUMNS definition
TRIPS_COLUMNS = {
    "Trip": [
        {"id": "distance", "name": "Distance", "source": "trip"},
        {"id": "drivingDuration", "name": "Driving Duration", "source": "trip"},
        {"id": "engineHours", "name": "Engine Hours", "source": "trip"},
        {"id": "idlingDuration", "name": "Idling Duration", "source": "trip"},
        {"id": "maximumSpeed", "name": "Maximum Speed", "source": "trip"},
        {"id": "nextTripStart", "name": "Next Trip Start", "source": "trip"},
        {"id": "start", "name": "Start Time", "source": "trip"},
        {"id": "stop", "name": "Stop Time", "source": "trip"},
        {"id": "stopDuration", "name": "Stop Duration", "source": "trip"},
        {"id": "workDistance", "name": "Work Distance", "source": "trip"},
        {"id": "workDrivingDuration", "name": "Work Driving Duration", "source": "trip"},
        {"id": "workStopDuration", "name": "Work Stop Duration", "source": "trip"}
    ],
    "Asset": AVAILABLE_COLUMNS["Asset"],  # Reuse existing device columns
    "Driver": AVAILABLE_COLUMNS["Driver"]  # Reuse existing user columns
}

@lru_cache(maxsize=None)
def decode_vin(vin: str) -> dict:
    """Decode VIN to get detailed vehicle information."""
    if not vin or vin == "":
        return {
            "make": "",
            "model": "",
            "year": "",
            "manufacturer": "",
            "plant": "",
            "vehicleType": "",
            "bodyStyle": "",
            "engineSize": "",
            "engineCylinders": "",
            "fuelType": "",
            "series": "",
            "transmission": "",
            "driveType": "",
            "gvwr": ""
        }
    
    url = f"https://vpic.nhtsa.dot.gov/api/vehicles/DecodeVin/{vin}?format=json"
    response = requests.get(url)
    
    if response.status_code != 200:
        return {
            "make": "",
            "model": "",
            "year": "",
            "manufacturer": "",
            "plant": "",
            "vehicleType": "",
            "bodyStyle": "",
            "engineSize": "",
            "engineCylinders": "",
            "fuelType": "",
            "series": "",
            "transmission": "",
            "driveType": "",
            "gvwr": ""
        }
        
    data = response.json()
    results = data.get("Results", [])
    
    # Create a mapping of variable names to their values
    vin_data = {
        item["Variable"].replace(" ", ""): item["Value"]
        for item in results
        if item["Value"] and item["Value"] != "Not Applicable"
    }
    
    return {
        "year": vin_data.get("ModelYear", ""),
        "make": vin_data.get("Make", ""),
        "model": vin_data.get("Model", ""),
        "manufacturer": vin_data.get("Manufacturer", ""),
        "plant": vin_data.get("PlantCountry", ""),
        "vehicleType": vin_data.get("VehicleType", ""),
        "bodyStyle": vin_data.get("BodyClass", ""),
        "engineSize": vin_data.get("DisplacementL", ""),
        "engineCylinders": vin_data.get("EngineCylinders", ""),
        "fuelType": vin_data.get("FuelTypePrimary", ""),
        "series": vin_data.get("Series", ""),
        "transmission": vin_data.get("TransmissionStyle", ""),
        "driveType": vin_data.get("DriveType", ""),
        "gvwr": vin_data.get("GVWR", "")
    }

# Add caching for API calls
@lru_cache(maxsize=None)
def get_geotab_data(type_name, **kwargs):
    """Cached wrapper for Geotab API calls"""
    api = initialize_geotab_api()
    return api.call("Get", typeName=type_name, **kwargs)

def authenticate_geotab():
    """Authenticate with the MyGeotab API."""
    username = GEO_USER
    database = GEO_DATABASE
    
    password = GEO_PASSWORD
        
    # Initialize MyGeotab API
    api = mygeotab.API(username=username, password=password, database=database)
    api.authenticate()
    print("Authenticated with Geotab successfully.")
    return api

def initialize_geotab_api():
    global geotab_api
    if geotab_api is None:
        geotab_api = authenticate_geotab()
    else:
        print("Using cached Geotab API instance.")
    return geotab_api

def convert_datetime(dt_value):
    """Convert datetime value to a user-friendly format."""
    if not dt_value or dt_value == "":
        return ""
    
    try:
        # If it's already a datetime object
        if isinstance(dt_value, (datetime.datetime, pd.Timestamp)):
            return dt_value.strftime('%Y-%m-%d %H:%M:%S')
            
        # If it's a string, parse it first
        if isinstance(dt_value, str):
            dt_obj = parser.parse(dt_value)
            return dt_obj.strftime('%Y-%m-%d %H:%M:%S')
            
        return ""
    except Exception as e:
        print(f"Error converting datetime {dt_value}: {e}")
        return ""

def format_duration(duration_value):
    """Format duration to a readable format (HH:MM:SS)."""
    if not duration_value or duration_value == "":
        return ""
    
    try:
        # If it's a string, parse it
        if isinstance(duration_value, str):
            # Handle format with days (e.g., "20.08:24:28.9220000")
            if '.' in duration_value:
                parts = duration_value.split('.')
                if len(parts) >= 2:
                    # Split into days and time
                    days = int(float(parts[0]))  # Convert to float first in case it has decimals
                    time_part = parts[1]
                    if ':' in time_part:
                        time_parts = time_part.split(':')
                        hours = int(time_parts[0]) + (days * 24)
                        minutes = int(time_parts[1])
                        seconds = int(float(time_parts[2]))  # Handle milliseconds
                        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Handle regular time format "HH:MM:SS" or "HH:MM:SS.ms"
            if ':' in duration_value:
                time_parts = duration_value.split(':')
                if len(time_parts) == 3:
                    hours = int(time_parts[0])
                    minutes = int(time_parts[1])
                    seconds = int(float(time_parts[2]))
                    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        
        # If it's a datetime.time object
        if hasattr(duration_value, 'hour'):
            return f"{duration_value.hour:02d}:{duration_value.minute:02d}:{duration_value.second:02d}"
        
        # If it's a number (seconds)
        if isinstance(duration_value, (int, float)):
            total_seconds = int(duration_value)
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
        return str(duration_value)
        
    except Exception as e:
        print(f"Error formatting duration {duration_value}: {e}")
        return str(duration_value)

def format_distance(distance_km):
    """Format distance in kilometers to miles with 2 decimal places."""
    try:
        # Handle None, empty string, or zero
        if distance_km is None or distance_km == "":
            return "0.00"
            
        # Convert to float if it's not already
        distance_km = float(distance_km)
        
        # Handle zero explicitly
        if distance_km == 0:
            return "0.00"
            
        # Format with 2 decimal places
        return f"{distance_km:.2f}"
        
    except Exception as e:
        print(f"Error formatting distance {distance_km}: {e}")
        return "0.00"


def generate_report_with_dates(from_date, to_date, custom_filename=None, selected_rules=None, selected_columns=None, selected_devices=None, sort_column=None, sort_ascending=True):
    """Generate report for specified date range, rules, columns, and devices."""
    global report_status, report_file_path
    
    if custom_filename:
        report_file_path = custom_filename
    else:
        # Generate filename with timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file_path = f"exception_report_{timestamp}.xlsx"
    
    try:
        report_status = {"status": "processing", "message": "Authenticating with Geotab..."}

        print("Authenticating with Geotab from exceptions")
        api = initialize_geotab_api()

        # Get all rules or filter by selected rules
        rules = get_geotab_data("Rule")
        if selected_rules:
            rules = [rule for rule in rules if rule['id'] in selected_rules]
        
        rule_map = {rule['id']: rule['name'] for rule in rules}

        report_status = {"status": "processing", "message": "Fetching exception data..."}
        
        # Fetch exceptions only for selected rules
        exceptions = []
        total_rules = len(rules)
        for i, rule in enumerate(rules, 1):
            report_status = {
                "status": "processing", 
                "message": f"Fetching exceptions for rule {i} of {total_rules}..."
            }
            
            search_criteria = {
                "fromDate": from_date,
                "toDate": to_date,
                "ruleSearch": {"id": rule['id']}
            }
            
            # Add device filter if devices are selected
            #if selected_devices:
            rule_exceptions = api.call("Get", typeName="ExceptionEvent", search=search_criteria)
            exceptions.extend([exc for exc in rule_exceptions if exc['device']['id'] in selected_devices])
            #else:
             #   rule_exceptions = api.call("Get", typeName="ExceptionEvent", search=search_criteria)  # Fixed line
              #  exceptions.extend(rule_exceptions)

        # Extract unique device IDs from exceptions
        device_ids = set(exception['device']['id'] for exception in exceptions)

        report_status = {"status": "processing", "message": "Fetching device data..."}
        
        # Fetch all devices in a single API call
        all_devices = get_geotab_data("Device")

        # Map device IDs to device details
        device_map = {device['id']: device for device in all_devices if device['id'] in device_ids}

        # Fetch DeviceStatusInfo in a single API call
        device_status_info = get_geotab_data("DeviceStatusInfo")

        # Map device IDs to DeviceStatusInfo
        device_status_map = {status['device']['id']: status for status in device_status_info}

        report_status = {"status": "processing", "message": "Fetching user data..."}
        
        # Fetch all users in a single API call
        all_users = get_geotab_data("User")

        # Map user IDs to user details
        user_map = {user['id']: user for user in all_users}

        # Fetch all groups in a single API call
        all_groups = get_geotab_data("Group")

        # Map group IDs to group names
        group_map = {group['id']: group['name'] for group in all_groups}

        report_status = {"status": "processing", "message": "Processing exception data..."}
        
        # Process each exception to gather required details
        processed_exceptions = []
        
        for exception in exceptions:
            row_data = {}
            
            for column in selected_columns:
                try:
                    # Handle column format safely
                    if '.' not in column:
                        print(f"Invalid column format: {column}")
                        continue
                        
                    source, field = column.split('.', 1)  # Split on first period only
                    
                    if source == 'exception':
                        if field == 'rule.id':
                            value = rule_map.get(exception.get('rule', {}).get('id', 'Unknown'), 'Unknown')
                        else:
                            value = exception.get(field, 'Unknown')
                            
                        # Format specific fields
                        if field == 'activeFrom' or field == 'activeTo':
                            value = convert_datetime(value)
                        elif field == 'duration':
                            value = format_duration(value)
                        elif field == 'distance':
                            value = format_distance(value)
                            
                    elif source == 'device':
                        device = device_map.get(exception['device']['id'], {})
                        # List of fields that should use VIN decoding
                        vin_fields = [
                            'make', 'model', 'year', 'manufacturer', 'plant', 'vehicleType',
                            'bodyStyle', 'engineSize', 'engineCylinders', 'fuelType', 'series',
                            'transmission', 'driveType', 'gvwr'
                        ]
                        
                        if field in vin_fields:
                            vin = device.get('vehicleIdentificationNumber')
                            if vin:
                                vin_details = decode_vin(vin)
                                value = vin_details.get(field, device.get(field, 'Unknown'))
                            else:
                                value = device.get(field, 'Unknown')
                        elif field == 'groups':
                            group_id = device.get('groups', [{}])[0].get('id', 'Unknown')
                            value = group_map.get(group_id, 'Unknown')
                        else:
                            value = device.get(field, 'Unknown')
                            
                    elif source == 'user':
                        driver = device_status_map.get(exception['device']['id'], {}).get('driver', {})
                        if not isinstance(driver, dict):
                            driver = {}
                        user = user_map.get(driver.get('id', 'Unknown'), {})
                        
                        if field == 'driverGroups':
                            group_id = user.get('driverGroups', [{}])[0].get('id', 'Unknown')
                            value = group_map.get(group_id, 'Unknown')
                        else:
                            value = user.get(field, 'Unknown')
                            
                    elif source == 'vin':
                        device = device_map.get(exception['device']['id'], {})
                        vin = device.get('vehicleIdentificationNumber', 'Unknown')
                        vin_details = decode_vin(vin)
                        value = vin_details.get(field, 'Unknown')
                    else:
                        print(f"Unknown source type: {source}")
                        continue
                        
                    row_data[f"{source}.{field}"] = value
                    
                except Exception as e:
                    print(f"Error processing column {column}: {e}")
                    row_data[column] = "Error"
                    continue
                    
            processed_exceptions.append(row_data)

        report_status = {"status": "processing", "message": "Creating Excel report..."}
        
        # Create DataFrame from processed exceptions
        df = pd.DataFrame(processed_exceptions)

        # Reorder columns based on selected_columns order
        available_columns = [col for col in selected_columns if col in df.columns]
        df = df[available_columns]
        
        # Get column display names directly from the dictionary
        #columns = get_available_columns()
        
        try:
            # Build column name mapping
            column_names = {}
            for group in AVAILABLE_COLUMNS:
                for col in AVAILABLE_COLUMNS[group]:
                    key = f"{col['source']}.{col['id']}"
                    column_names[key] = col['name']
            
            # Convert DataFrame column names
            existing_columns = set(df.columns)
            rename_dict = {col: column_names.get(col, col) for col in existing_columns}
            
            # Handle datetime columns before renaming
            for col in df.columns:
                if any(col.endswith(field) for field in ['activeFrom', 'activeTo', 'lastModifiedDateTime']):
                    print(f"\nProcessing datetime column: {col}")
                    try:
                        # Convert string timestamps to datetime objects and remove timezone
                        df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)
                        # Format to string
                        df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                        print(f"Successfully converted {col}")
                    except Exception as e:
                        print(f"Error converting column {col}: {e}")
                        continue
            
            # Rename columns after datetime conversion
            df.rename(columns=rename_dict, inplace=True)
            
        except Exception as e:
            print(f"Error processing columns: {e}")
            traceback.print_exc()
            # Continue with original column names if renaming fails
            pass

        # After creating DataFrame and before writing to Excel
        if sort_column:
            try:
                # Convert the sort_column to the display name if it exists
                display_name = None
                for group in AVAILABLE_COLUMNS:
                    for col in AVAILABLE_COLUMNS[group]:
                        if f"{col['source']}.{col['id']}" == sort_column:
                            display_name = col['name']
                            break
                    if display_name:
                        break

                if display_name and display_name in df.columns:
                    # Get sample value
                    sample_value = df[display_name].dropna().iloc[0] if not df[display_name].dropna().empty else None

                    if 'Distance' in display_name:  # Check if it's a distance column
                        # Convert to float explicitly for correct numeric sorting
                        df[display_name] = df[display_name].apply(lambda x: float(x) if x not in ['', 'Unknown'] else 0)
                        df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)  # Use sort_ascending parameter
                        # Convert back to string format after sorting
                        df[display_name] = df[display_name].apply(lambda x: f"{float(x):.2f}")
                    elif pd.api.types.is_numeric_dtype(df[display_name]):
                        df[display_name] = pd.to_numeric(df[display_name], errors='coerce')
                        df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)  # Use sort_ascending parameter
                    elif display_name in ['Start Time', 'Stop Time', 'Last Modified']:
                        # Try datetime first
                        try:
                            df[display_name] = pd.to_datetime(df[display_name], format='%Y-%m-%d %H:%M:%S', errors='coerce')
                            df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)
                        except (ValueError, TypeError):
                            # If not datetime, sort as strings
                            df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)  
                    else:
                        # Sort as strings
                        df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)  
                        print("Sorted by string")  
                    print(f"Sorted by {display_name}")
            except Exception as e:
                print(f"Error sorting by column {sort_column}: {e}")
                traceback.print_exc()

        try:
            # Generate Excel file
            with pd.ExcelWriter(report_file_path, engine='openpyxl') as writer:
                # Write the DataFrame without index
                df.to_excel(writer, index=False, sheet_name="Exception Report")
                
                # Get the worksheet
                worksheet = writer.sheets["Exception Report"]
                
                # Auto-adjust column widths
                for i, col in enumerate(df.columns):
                    # Get maximum length of column data
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    ) + 2  # Add a little extra space
                    
                    # Convert column number to letter (0 = A, 1 = B, etc.)
                    column_letter = get_column_letter(i + 1)  # Add 1 because openpyxl uses 1-based indexing
                    
                    # Set column width (max 50 characters)
                    worksheet.column_dimensions[column_letter].width = min(max_length, 50)

                # Add report metadata
                metadata_row = 1
                metadata_col = len(df.columns) + 2  # Two columns after the last data column
                
                # Add report generation details
                worksheet.cell(row=metadata_row, column=metadata_col, value="Report Details")
                worksheet.cell(row=metadata_row + 1, column=metadata_col, value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                worksheet.cell(row=metadata_row + 2, column=metadata_col, value=f"Date Range: {from_date[:10]} to {to_date[:10]}")
                worksheet.cell(row=metadata_row + 3, column=metadata_col, value=f"Total Records: {len(df)}")
                
                # Style the metadata
                for row in range(metadata_row, metadata_row + 4):
                    cell = worksheet.cell(row=row, column=metadata_col)
                    cell.font = Font(bold=True if row == metadata_row else False)
                
            report_status = {"status": "complete", "message": "Report generation complete"}
            print(f"Report generated successfully: {report_file_path}")
            
        except Exception as e:
            error_msg = f"Error generating Excel file: {str(e)}"
            print(error_msg)
            report_status = {"status": "error", "message": error_msg}
            raise
            
        return len(processed_exceptions)
        
    except Exception as e:
        error_msg = f"Error in report generation: {str(e)}"
        print(error_msg)
        print("Full traceback:")
        traceback.print_exc()  # Print full stack trace
        print(f"Selected devices: {selected_devices}")  # Debug selected devices
        print(f"From date: {from_date}")  # Debug dates
        print(f"To date: {to_date}")
        report_status = {"status": "error", "message": error_msg}
        return 0

def generate_trips_report(from_date, to_date, selected_columns=None, selected_devices=None, sort_column=None, sort_ascending=True):  # Add sort_column parameter
    """Generate trips report for specified date range, columns, and devices."""
    global report_status, report_file_path
    
    # Generate filename with timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file_path = f"trips_report_{timestamp}.xlsx"
    
    try:
        report_status = {"status": "processing", "message": "Authenticating with Geotab..."}
        
        print("Authenticating with Geotab from trips")
        # Initialize API connection
        api = initialize_geotab_api()
        
        report_status = {"status": "processing", "message": "Fetching trip data..."}
        
        # Fetch trips for selected devices
        trips = []
        total_devices = len(selected_devices) if selected_devices else 0
        
        if selected_devices and total_devices > 0:
            for i, device_id in enumerate(selected_devices, 1):
                report_status = {
                    "status": "processing",
                    "message": f"Fetching trips for device {i} of {total_devices}..."
                }
                
                # Create device search object
                device_search = {"id": device_id}
                
                # Fetch trips for this device in the date range
                device_trips = api.call("Get", typeName="Trip", 
                                     search={
                                         "fromDate": from_date,
                                         "toDate": to_date,
                                         "deviceSearch": device_search
                                     })
                
                trips.extend(device_trips)
        else:
            # Fetch all trips in the date range
            trips = api.call("Get", typeName="Trip", 
                          search={
                              "fromDate": from_date,
                              "toDate": to_date
                          })
            
        # Extract unique device IDs from trips
        device_ids = set()
        for trip in trips:
            if 'device' in trip:
                if isinstance(trip['device'], dict) and 'id' in trip['device']:
                    device_ids.add(trip['device']['id'])
                elif isinstance(trip['device'], str):
                    device_ids.add(trip['device'])
        
        report_status = {"status": "processing", "message": "Fetching device data..."}
        
        # Fetch all devices in a single API call
        all_devices = get_geotab_data("Device")
        
        # Map device IDs to device details
        device_map = {device['id']: device for device in all_devices if device['id'] in device_ids}
        
        # Fetch DeviceStatusInfo in a single API call
        device_status_info = get_geotab_data("DeviceStatusInfo")
        
        # Map device IDs to DeviceStatusInfo
        device_status_map = {status['device']['id']: status for status in device_status_info 
                            if isinstance(status['device'], dict) and 'id' in status['device']}
        
        # Fetch all users in a single API call
        all_users = get_geotab_data("User")
        
        # Map user IDs to user details
        user_map = {user['id']: user for user in all_users if 'id' in user}
        
        # Fetch all groups in a single API call
        all_groups = get_geotab_data("Group")
        
        # Map group IDs to group names
        group_map = {group['id']: group['name'] for group in all_groups if 'id' in group}
        
        report_status = {"status": "processing", "message": "Processing trip data..."}
        
        # Process each trip to gather required details
        processed_trips = []
        
        for trip in trips:
            row_data = {}
            
            for column in selected_columns:
                try:
                    if '.' not in column:
                        print(f"Invalid column format: {column}")
                        continue
                        
                    source, field = column.split('.', 1)
                    
                    if source == 'trip':
                        value = trip.get(field, 'Unknown')
                        
                        # Format specific fields
                        if field in ['start', 'stop', 'nextTripStart']:
                            value = convert_datetime(value)
                        elif field in ['drivingDuration', 'idlingDuration', 'stopDuration', 
                                     'workDrivingDuration', 'workStopDuration']:
                            value = format_duration(value)
                        elif field in ['distance', 'workDistance']:
                            value = format_distance(value)
                    
                    elif source == 'device':
                        # Get device ID properly from trip
                        device_id = None
                        if isinstance(trip.get('device'), dict) and 'id' in trip['device']:
                            device_id = trip['device']['id']
                        elif isinstance(trip.get('device'), str):
                            device_id = trip['device']
                            
                        if not device_id or device_id not in device_map:
                            value = 'Unknown'
                            continue
                            
                        device = device_map.get(device_id, {})
                        
                        # Handle VIN-related fields
                        if field in ['make', 'model', 'year', 'manufacturer', 'plant', 'vehicleType',
                                  'bodyStyle', 'engineSize', 'engineCylinders', 'fuelType', 'series',
                                  'transmission', 'driveType', 'gvwr']:
                            vin = device.get('vehicleIdentificationNumber', '')
                            if vin:
                                vin_details = decode_vin(vin)
                                value = vin_details.get(field, device.get(field, 'Unknown'))
                            else:
                                value = device.get(field, 'Unknown')
                        elif field == 'groups':
                            groups = device.get('groups', [])
                            if groups and len(groups) > 0 and isinstance(groups[0], dict) and 'id' in groups[0]:
                                group_id = groups[0]['id']
                                value = group_map.get(group_id, 'Unknown')
                            else:
                                value = 'Unknown'
                        else:
                            value = device.get(field, 'Unknown')
                    
                    elif source == 'user':
                        # Get device ID properly
                        device_id = None
                        if isinstance(trip.get('device'), dict) and 'id' in trip['device']:
                            device_id = trip['device']['id']
                        elif isinstance(trip.get('device'), str):
                            device_id = trip['device']
                            
                        if not device_id or device_id not in device_status_map:
                            value = 'Unknown'
                            continue
                            
                        status = device_status_map.get(device_id, {})
                        driver = status.get('driver', {})
                        
                        if not driver or not isinstance(driver, dict) or 'id' not in driver:
                            value = 'Unknown'
                            continue
                            
                        user_id = driver.get('id')
                        if not user_id or user_id not in user_map:
                            value = 'Unknown'
                            continue
                            
                        user = user_map.get(user_id, {})
                        
                        if field == 'driverGroups':
                            # Check both companyGroups and driverGroups
                            groups = user.get('companyGroups', []) or user.get('driverGroups', [])
                            if groups and len(groups) > 0 and isinstance(groups[0], dict) and 'id' in groups[0]:
                                group_id = groups[0]['id']
                                value = group_map.get(group_id, 'Unknown')
                            else:
                                value = 'Unknown'
                        else:
                            value = user.get(field, 'Unknown')
                    
                    row_data[f"{source}.{field}"] = value
                    
                except Exception as e:
                    print(f"Error processing column {column}: {e}")
                    traceback.print_exc()
                    row_data[column] = "Error"
                    continue
                    
            processed_trips.append(row_data)

        # Only continue if we have data
        if not processed_trips:
            report_status = {"status": "error", "message": "No trip data found for the selected criteria"}
            return 0
            
        report_status = {"status": "processing", "message": "Creating Excel report..."}
        
        # Create DataFrame from processed trips
        df = pd.DataFrame(processed_trips)

        # Handle empty dataframe gracefully
        if df.empty:
            report_status = {"status": "error", "message": "No data available for the report"}
            return 0
            
        # Use only columns that exist in the dataframe
        available_columns = [col for col in selected_columns if col in df.columns]
        if available_columns:
            df = df[available_columns]
        
        try:
            # Build column name mapping
            column_names = {}
            for group in TRIPS_COLUMNS:
                for col in TRIPS_COLUMNS[group]:
                    key = f"{col['source']}.{col['id']}"
                    column_names[key] = col['name']
            
            # Rename columns
            df.rename(columns=column_names, inplace=True)
            
            # After creating DataFrame and before writing to Excel
            if sort_column:
                try:
                    # Convert the sort_column to the display name if it exists
                    display_name = None
                    for group in TRIPS_COLUMNS:
                        for col in TRIPS_COLUMNS[group]:
                            if f"{col['source']}.{col['id']}" == sort_column:
                                display_name = col['name']
                                break
                        if display_name:
                            break

                    if display_name and display_name in df.columns:
                        # Get sample value
                        sample_value = df[display_name].dropna().iloc[0] if not df[display_name].dropna().empty else None

                        if 'Distance' or 'Engine Hours' in display_name:  # Check if it's a distance column
                            # Convert to float explicitly for correct numeric sorting
                            df[display_name] = df[display_name].apply(lambda x: float(x) if x not in ['', 'Unknown'] else 0)
                            df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)  # Use sort_ascending parameter
                            # Convert back to string format after sorting
                            df[display_name] = df[display_name].apply(lambda x: f"{float(x):.2f}")
                            print("Sorted numerically")
                        elif pd.api.types.is_numeric_dtype(df[display_name]):
                            df[display_name] = pd.to_numeric(df[display_name], errors='coerce')
                            df.sort_values(by=display_name, ascending=sort_ascending, inplace=True) 
                            print("Sorted numerically")
                        elif display_name in ['Start Time', 'Stop Time', 'Next Trip Start']:
                            # Try datetime first
                            try:
                                df[display_name] = pd.to_datetime(df[display_name], format='%Y-%m-%d %H:%M:%S', errors='coerce')
                                df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)
                                print("Sorted by datetime")
                            except (ValueError, TypeError):
                                # If not datetime, sort as strings
                                df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)  
                                print("Sorted by string")
                        else:
                            # Sort as strings
                            df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)  
                            print("Sorted by string")
                        print("I am here")
                        print(f"Sorted by RIGHT HERE{display_name}")
                    else:
                        print(f"Warning: Column '{display_name}' not found in DataFrame. Skipping sort.")
                except Exception as e:
                    print(f"Error sorting by column {sort_column}: {e}")
                    traceback.print_exc()

            # Generate Excel file
            with pd.ExcelWriter(report_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Trips Report")
                
                worksheet = writer.sheets["Trips Report"]
                
                # Auto-adjust column widths
                for i, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    ) + 2
                    
                    column_letter = get_column_letter(i + 1)
                    worksheet.column_dimensions[column_letter].width = min(max_length, 50)

                # Add report metadata
                metadata_row = 1
                metadata_col = len(df.columns) + 2
                
                worksheet.cell(row=metadata_row, column=metadata_col, value="Report Details")
                worksheet.cell(row=metadata_row + 1, column=metadata_col, value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                worksheet.cell(row=metadata_row + 2, column=metadata_col, value=f"Date Range: {from_date[:10]} to {to_date[:10]}")
                worksheet.cell(row=metadata_row + 3, column=metadata_col, value=f"Total Records: {len(df)}")
                
                for row in range(metadata_row, metadata_row + 4):
                    cell = worksheet.cell(row=row, column=metadata_col)
                    cell.font = Font(bold=True if row == metadata_row else False)
                
            report_status = {"status": "complete", "message": "Report generation complete"}
            print(f"Report generated successfully: {report_file_path}")
            return len(processed_trips)
            
        except Exception as e:
            error_msg = f"Error generating Excel file: {str(e)}"
            print(error_msg)
            report_status = {"status": "error", "message": error_msg}
            raise
            
    except Exception as e:
        error_msg = f"Error in report generation: {str(e)}"
        print(error_msg)
        report_status = {"status": "error", "message": error_msg}
        return 0

def generate_faults_report(from_date, to_date, custom_filename=None, selected_diagnostics=None, selected_columns=None, selected_devices=None, sort_column=None, sort_ascending=True):
    """Generate fault report for specified date range, diagnostics, columns, and devices."""
    global report_status, report_file_path

    print("Generating fault report...")
    
    if custom_filename:
        report_file_path = custom_filename
    else:
        # Generate filename with timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file_path = f"fault_report_{timestamp}.xlsx"
    
    try:
        report_status = {"status": "processing", "message": "Authenticating with Geotab..."}

        print("Authenticating with Geotab from faults")
        api = initialize_geotab_api()

        # Only get diagnostic data for selected diagnostics if they are provided
        diagnostics = []
        if selected_diagnostics:
            report_status = {"status": "processing", "message": "Fetching selected diagnostics..."}
            # Fetch only the selected diagnostics
            count = 0
            for diagnostic_id in selected_diagnostics:
                count += 1
                print(f"Fetching diagnostic {count} of {len(selected_diagnostics)}...")
                search_criteria = {"id": diagnostic_id}
                diagnostic_data = api.call("Get", typeName="Diagnostic", search=search_criteria)
                if diagnostic_data:
                    diagnostics.extend(diagnostic_data)
        
        diagnostic_map = {diagnostic['id']: diagnostic for diagnostic in diagnostics}

        report_status = {"status": "processing", "message": "Fetching fault data..."}
        
        # Fetch faults only for selected diagnostics
        faults = []
        total_diagnostics = len(diagnostics)
        for i, diagnostic in enumerate(diagnostics, 1):
            report_status = {
                "status": "processing", 
                "message": f"Fetching faults for diagnostic {i} of {total_diagnostics}..."
            }
            
            search_criteria = {
                "fromDate": from_date,
                "toDate": to_date,
                "diagnosticSearch": {"id": diagnostic['id']}
            }
            
            diagnostic_faults = api.call("Get", typeName="FaultData", search=search_criteria)
            # Add device filter if devices are selected
            if selected_devices:
                diagnostic_faults = [fault for fault in diagnostic_faults if fault['device']['id'] in selected_devices]
            
            faults.extend(diagnostic_faults)

        # Extract unique device IDs from faults
        device_ids = set(fault['device']['id'] for fault in faults)

        report_status = {"status": "processing", "message": "Fetching device data..."}
        
        # Fetch all devices in a single API call
        all_devices = get_geotab_data("Device")

        # Map device IDs to device details
        device_map = {device['id']: device for device in all_devices if device['id'] in device_ids}

        # Fetch DeviceStatusInfo in a single API call to get driver information
        device_status_info = get_geotab_data("DeviceStatusInfo")

        # Map device IDs to DeviceStatusInfo
        device_status_map = {status['device']['id']: status for status in device_status_info}

        report_status = {"status": "processing", "message": "Fetching user data..."}
        
        # Fetch all users in a single API call
        all_users = get_geotab_data("User")

        # Map user IDs to user details
        user_map = {user['id']: user for user in all_users}

        # Fetch all groups in a single API call
        all_groups = get_geotab_data("Group")

        # Map group IDs to group names
        group_map = {group['id']: group['name'] for group in all_groups}

        report_status = {"status": "processing", "message": "Processing fault data..."}

        print(f"Processing {len(faults)} faults...")
        
        # Process each fault to gather required details
        processed_faults = []
        
        for fault in faults:
            row_data = {}
            
            for column in selected_columns:
                try:
                    # Handle column format safely
                    if '.' not in column:
                        print(f"Invalid column format: {column}")
                        continue
                        
                    source, field = column.split('.', 1)  # Split on first period only
                    
                    # Handle the mapped sources according to FAULTS_COLUMNS
                    if source == 'faultdata':
                        # Map to the internal fault object
                        if field == 'dateTime':
                            value = convert_datetime(fault.get(field, 'Unknown'))
                        else:
                            value = fault.get(field, 'Unknown')
                            
                    elif source == 'diagnostic':
                        # Get diagnostic data from the diagnostic_map
                        diagnostic_id = fault.get('diagnostic', {}).get('id')
                        if diagnostic_id and diagnostic_id in diagnostic_map:
                            diagnostic_obj = diagnostic_map[diagnostic_id]
                            if field == 'name':
                                value = diagnostic_obj.get('name', 'Unknown')
                            elif field == 'faultCode':
                                value = diagnostic_obj.get('code', 'Unknown')
                            else:
                                value = diagnostic_obj.get(field, 'Unknown')
                        else:
                            value = 'Unknown'
                            
                    elif source == 'device':
                        device = device_map.get(fault['device']['id'], {})
                        # List of fields that should use VIN decoding
                        vin_fields = [
                            'make', 'model', 'year', 'manufacturer', 'plant', 'vehicleType',
                            'bodyStyle', 'engineSize', 'engineCylinders', 'fuelType', 'series',
                            'transmission', 'driveType', 'gvwr'
                        ]
                        
                        if field in vin_fields:
                            vin = device.get('vehicleIdentificationNumber')
                            if vin:
                                vin_details = decode_vin(vin)
                                value = vin_details.get(field, device.get(field, 'Unknown'))
                            else:
                                value = device.get(field, 'Unknown')
                        elif field == 'groups':
                            group_id = device.get('groups', [{}])[0].get('id', 'Unknown')
                            value = group_map.get(group_id, 'Unknown')
                        else:
                            value = device.get(field, 'Unknown')
                            
                    elif source == 'user':
                        # For FaultData, we need to get driver info from DeviceStatusInfo
                        device_status = device_status_map.get(fault['device']['id'], {})
                        driver = device_status.get('driver', {})
                        if not isinstance(driver, dict):
                            driver = {}
                        user = user_map.get(driver.get('id', 'Unknown'), {})
                        
                        if field == 'driverGroups':
                            group_id = user.get('driverGroups', [{}])[0].get('id', 'Unknown')
                            value = group_map.get(group_id, 'Unknown')
                        else:
                            value = user.get(field, 'Unknown')
                            
                    elif source == 'vin':
                        device = device_map.get(fault['device']['id'], {})
                        vin = device.get('vehicleIdentificationNumber', 'Unknown')
                        vin_details = decode_vin(vin)
                        value = vin_details.get(field, 'Unknown')
                    else:
                        print(f"Unknown source type: {source}")
                        continue
                        
                    row_data[f"{source}.{field}"] = value
                    
                except Exception as e:
                    print(f"Error processing column {column}: {e}")
                    row_data[column] = "Error"
                    continue
                    
            processed_faults.append(row_data)

        report_status = {"status": "processing", "message": "Creating Excel report..."}
        
        # Create DataFrame from processed faults
        df = pd.DataFrame(processed_faults)

        # Reorder columns based on selected_columns order
        available_columns = [col for col in selected_columns if col in df.columns]
        df = df[available_columns]
        
        try:
            # Build column name mapping
            column_names = {}
            for group in FAULTS_COLUMNS:
                for col in FAULTS_COLUMNS[group]:
                    key = f"{col['source']}.{col['id']}"
                    column_names[key] = col['name']
            
            # Convert DataFrame column names
            existing_columns = set(df.columns)
            rename_dict = {col: column_names.get(col, col) for col in existing_columns}
            
            # Handle datetime columns before renaming
            for col in df.columns:
                if any(col.endswith(field) for field in ['dateTime', 'lastModifiedDateTime']):
                    print(f"\nProcessing datetime column: {col}")
                    try:
                        # Convert string timestamps to datetime objects and remove timezone
                        df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)
                        # Format to string
                        df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                        print(f"Successfully converted {col}")
                    except Exception as e:
                        print(f"Error converting column {col}: {e}")
                        continue
            
            # Rename columns after datetime conversion
            df.rename(columns=rename_dict, inplace=True)
            
        except Exception as e:
            print(f"Error processing columns: {e}")
            traceback.print_exc()
            # Continue with original column names if renaming fails
            pass

        # After creating DataFrame and before writing to Excel
        if sort_column:
            try:
                # Convert the sort_column to the display name if it exists
                display_name = None
                for group in FAULTS_COLUMNS:
                    for col in FAULTS_COLUMNS[group]:
                        if f"{col['source']}.{col['id']}" == sort_column:
                            display_name = col['name']
                            break
                    if display_name:
                        break

                if display_name and display_name in df.columns:
                    # Get sample value
                    sample_value = df[display_name].dropna().iloc[0] if not df[display_name].dropna().empty else None

                    if 'Distance' in display_name:  # Check if it's a distance column
                        # Convert to float explicitly for correct numeric sorting
                        df[display_name] = df[display_name].apply(lambda x: float(x) if x not in ['', 'Unknown'] else 0)
                        df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)
                        # Convert back to string format after sorting
                        df[display_name] = df[display_name].apply(lambda x: f"{float(x):.2f}")
                    elif pd.api.types.is_numeric_dtype(df[display_name]):
                        df[display_name] = pd.to_numeric(df[display_name], errors='coerce')
                        df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)
                    elif display_name in ['Date Time', 'Last Modified']:
                        # Try datetime first
                        try:
                            df[display_name] = pd.to_datetime(df[display_name], format='%Y-%m-%d %H:%M:%S', errors='coerce')
                            df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)
                        except (ValueError, TypeError):
                            # If not datetime, sort as strings
                            df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)
                    else:
                        # Sort as strings
                        df.sort_values(by=display_name, ascending=sort_ascending, inplace=True)
                    print(f"Sorted by {display_name}")
            except Exception as e:
                print(f"Error sorting by column {sort_column}: {e}")
                traceback.print_exc()

        try:
            # Generate Excel file
            with pd.ExcelWriter(report_file_path, engine='openpyxl') as writer:
                # Write the DataFrame without index
                df.to_excel(writer, index=False, sheet_name="Fault Report")
                
                # Get the worksheet
                worksheet = writer.sheets["Fault Report"]
                
                # Auto-adjust column widths
                for i, col in enumerate(df.columns):
                    # Get maximum length of column data
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    ) + 2  # Add a little extra space
                    
                    # Convert column number to letter (0 = A, 1 = B, etc.)
                    column_letter = get_column_letter(i + 1)  # Add 1 because openpyxl uses 1-based indexing
                    
                    # Set column width (max 50 characters)
                    worksheet.column_dimensions[column_letter].width = min(max_length, 50)

                # Add report metadata
                metadata_row = 1
                metadata_col = len(df.columns) + 2  # Two columns after the last data column
                
                # Add report generation details
                worksheet.cell(row=metadata_row, column=metadata_col, value="Report Details")
                worksheet.cell(row=metadata_row + 1, column=metadata_col, value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                worksheet.cell(row=metadata_row + 2, column=metadata_col, value=f"Date Range: {from_date[:10]} to {to_date[:10]}")
                worksheet.cell(row=metadata_row + 3, column=metadata_col, value=f"Total Records: {len(df)}")
                
                # Style the metadata
                for row in range(metadata_row, metadata_row + 4):
                    cell = worksheet.cell(row=row, column=metadata_col)
                    cell.font = Font(bold=True if row == metadata_row else False)
                
            report_status = {"status": "complete", "message": "Report generation complete"}
            print(f"Report generated successfully: {report_file_path}")
            
        except Exception as e:
            error_msg = f"Error generating Excel file: {str(e)}"
            print(error_msg)
            report_status = {"status": "error", "message": error_msg}
            raise
            
        return len(processed_faults)
        
    except Exception as e:
        error_msg = f"Error in report generation: {str(e)}"
        print(error_msg)
        print("Full traceback:")
        traceback.print_exc()  # Print full stack trace
        print(f"Selected devices: {selected_devices}")  # Debug selected devices
        print(f"Selected diagnostics: {selected_diagnostics}")  # Debug selected diagnostics
        print(f"From date: {from_date}")  # Debug dates
        print(f"To date: {to_date}")
        report_status = {"status": "error", "message": error_msg}
        return 0

@app.route('/')
def home():
    return send_file('ACreport.html')

@app.route('/build-report')
def build_report_page():
    return send_file('build-report.html')

@app.route('/build-report.js')
def build_report_js():
    return send_file('build-report.js')

@app.route('/build-trips-report')
def build_trips_report_page():
    return send_file('trips-report.html')

@app.route('/trips-report.js')
def trips_report_js():
    return send_file('trips-report.js')

@app.route('/build-faults-report')
def build_faults_report_page():
    return send_file('faults-report.html')

@app.route('/faults-report.js')
def faults_report_js():
    return send_file('faults-report.js')

@app.route('/api/rules', methods=['GET'])
def get_rules():
    """Endpoint to fetch available rules from Geotab."""
    try:
        rules = get_geotab_data("Rule")
        
        # Filter out "Application Exception" rule and format rules for frontend
        formatted_rules = [{
            'id': rule['id'],
            'name': rule['name'],
            'description': rule.get('description', '')
        } for rule in rules if rule['name'] != "**Application Exception"]
        
        # Sort alphabetically by name
        formatted_rules.sort(key=lambda x: x['name'].lower())
        
        return jsonify(formatted_rules)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/api/diagnostics', methods=['GET'])
def get_diagnostics():
    """Endpoint to fetch available rules from Geotab."""
    try:
        api = initialize_geotab_api()

        search_criteria = {
            "diagnosticType" : "GoFault"
        }

        diagnostics = api.call("Get", typeName="Diagnostic", search=search_criteria)
        
        # Filter out "Application Exception" rule and format rules for frontend
        formatted_diagnostics = [{
            'id': diagnostic['id'],
            'name': diagnostic['name']
        } for diagnostic in diagnostics]
        
        # Sort alphabetically by name
        formatted_diagnostics.sort(key=lambda x: x['name'].lower())
        
        return jsonify(formatted_diagnostics)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Add new endpoint to get available columns
@app.route('/api/available-columns', methods=['GET'])
def get_available_columns():
    """Get all available columns from different data sources."""
    try:
        # Create a copy of the columns to avoid modifying the original
        columns = {
            group: sorted(cols, key=lambda x: x['name'].lower())
            for group, cols in AVAILABLE_COLUMNS.items()
        }
        return jsonify(columns)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/api/available-faults-columns', methods=['GET'])
def get_faults_columns():
    """Get all available columns from different data sources."""
    try:
        # Create a copy of the columns to avoid modifying the original
        columns = {
            group: sorted(cols, key=lambda x: x['name'].lower())
            for group, cols in FAULTS_COLUMNS.items()
        }
        return jsonify(columns)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/generate-report', methods=['POST'])
def generate_report():
    """Endpoint to generate report with selected rules and columns."""
    global report_status
    
    try:
        data = request.get_json()
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        rule_ids = data.get('ruleIds', [])
        selected_columns = data.get('columns', [])
        selected_devices = data.get('deviceIds', [])  # Add device IDs
        sort_column = data.get('sortColumn')
        sort_ascending = data.get('sortAscending', True)  # Add sort_ascending parameter

        print(f"Selected devices: {selected_devices}")  # Debug selected devices
        
        if not all([from_date, to_date, rule_ids, selected_columns]):
            return jsonify({
                "status": "error", 
                "message": "Missing required parameters"
            }), 400

        # Ensure dates are in UTC
        try:
            from_date = parser.parse(from_date).astimezone(datetime.timezone.utc).isoformat()
            to_date = parser.parse(to_date).astimezone(datetime.timezone.utc).isoformat()
        except Exception as e:
            return jsonify({
                "status": "error",
                "message": f"Invalid date format: {str(e)}"
            }), 400

        report_status = {
            "status": "processing", 
            "message": "Starting report generation..."
        }
        
        thread = Thread(
            target=generate_report_with_dates, 
            args=(from_date, to_date),
            kwargs={
                "selected_rules": rule_ids,
                "selected_columns": selected_columns,
                "selected_devices": selected_devices,  # Add selected devices
                "sort_column": sort_column,  # Add sort column
                "sort_ascending": sort_ascending  # Add sort_ascending parameter
            }
        )
        thread.daemon = True
        thread.start()
        
        return jsonify(report_status)
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    
@app.route('/api/trips-columns', methods=['GET'])
def get_trips_columns():
    """Get all available columns for trips report."""
    try:
        columns = {
            group: sorted(cols, key=lambda x: x['name'].lower())
            for group, cols in TRIPS_COLUMNS.items()
        }
        return jsonify(columns)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/generate-trips-report', methods=['POST'])
def generate_trips_report_endpoint():
    """Endpoint to generate trips report."""
    global report_status
    
    try:
        data = request.get_json()
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        selected_columns = data.get('columns', [])
        selected_devices = data.get('deviceIds', [])
        sort_column = data.get('sortColumn')
        sort_ascending = data.get('sortAscending', True)  # Add sort_ascending parameter
        
        if not all([from_date, to_date, selected_columns]):
            return jsonify({
                "status": "error", 
                "message": "Missing required parameters"
            }), 400

        # Convert dates to UTC
        try:
            from_date = parser.parse(from_date).astimezone(datetime.timezone.utc).isoformat()
            to_date = parser.parse(to_date).astimezone(datetime.timezone.utc).isoformat()
        except Exception as e:
            return jsonify({
                "status": "error",
                "message": f"Invalid date format: {str(e)}"
            }), 400

        # Set initial report status
        report_status = {
            "status": "processing", 
            "message": "Starting trip report generation..."
        }
        
        thread = Thread(
            target=generate_trips_report,
            args=(from_date, to_date),
            kwargs={
                "selected_columns": selected_columns,
                "selected_devices": selected_devices,
                "sort_column": sort_column,  # Add sort column
                "sort_ascending": sort_ascending  # Add sort_ascending parameter
            }
        )
        thread.daemon = True
        thread.start()
        
        return jsonify(report_status)
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    
@app.route('/api/generate-faults-report', methods=['POST'])
def generate_faults_report_endpoint():
    """Endpoint to generate report with selected rules and columns."""
    global report_status
    
    try:
        data = request.get_json()
        from_date = data.get('fromDate')
        to_date = data.get('toDate')
        diagnostic_ids = data.get('diagnosticIds', [])
        selected_columns = data.get('columns', [])
        selected_devices = data.get('deviceIds', [])  # Add device IDs
        sort_column = data.get('sortColumn')
        sort_ascending = data.get('sortAscending', True)  # Add sort_ascending parameter
        
        if not all([from_date, to_date, diagnostic_ids, selected_columns]):
            return jsonify({
                "status": "error", 
                "message": "Missing required parameters"
            }), 400

        # Ensure dates are in UTC
        try:
            from_date = parser.parse(from_date).astimezone(datetime.timezone.utc).isoformat()
            to_date = parser.parse(to_date).astimezone(datetime.timezone.utc).isoformat()
        except Exception as e:
            return jsonify({
                "status": "error",
                "message": f"Invalid date format: {str(e)}"
            }), 400

        report_status = {
            "status": "processing", 
            "message": "Starting report generation..."
        }
        
        thread = Thread(
            target=generate_faults_report, 
            args=(from_date, to_date),
            kwargs={
                "selected_diagnostics": diagnostic_ids,
                "selected_columns": selected_columns,
                "selected_devices": selected_devices,  # Add selected devices
                "sort_column": sort_column,  # Add sort column
                "sort_ascending": sort_ascending  # Add sort_ascending parameter
            }
        )
        thread.daemon = True
        thread.start()
        
        return jsonify(report_status)
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

# Add a new endpoint to get all driver groups
@app.route('/api/driver-groups', methods=['GET'])
def get_driver_groups():
    """Endpoint to fetch available driver groups from Geotab."""
    try:
        groups = get_geotab_data("Group")
        
        # Filter only groups that are likely driver groups (can adjust this logic if needed)
        driver_groups = [{
            'id': group['id'],
            'name': group['name']
        } for group in groups]
        
        # Sort alphabetically by name
        driver_groups.sort(key=lambda x: x['name'].lower())
        
        return jsonify(driver_groups)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Modify the devices endpoint to accept driver group filters
@app.route('/api/devices', methods=['GET'])
def get_devices():
    """Endpoint to fetch available devices from Geotab, with optional driver group filtering."""
    try:
        driver_group_ids = request.args.getlist('driverGroupIds[]')
        
        devices = get_geotab_data("Device")
        
        # Get all device status info to find drivers
        device_status_info = get_geotab_data("DeviceStatusInfo")
        
        # Create device to driver mapping - handle both dictionary and string formats
        device_to_driver = {}
        for status in device_status_info:
            if not status.get('device') or not status.get('driver'):
                continue
                
            # Handle device ID properly whether it's a dict or string
            if isinstance(status['device'], dict):
                device_id = status['device'].get('id')
            else:
                device_id = status['device']
                
            # Handle driver ID properly whether it's a dict or string
            if isinstance(status['driver'], dict):
                driver_id = status['driver'].get('id')
            else:
                driver_id = status['driver']
                
            if device_id and driver_id:
                device_to_driver[device_id] = driver_id
        
        # If driver groups selected, get all users to filter by group
        if driver_group_ids:
            # Get all users
            users = get_geotab_data("User")
            
            # Create set of user IDs that belong to selected driver groups
            filtered_user_ids = set()
            for user in users:
                if not user.get('id'):
                    continue
                    
                user_groups = user.get('companyGroups', [])
                if not user_groups:
                    continue
                    
                for group in user_groups:
                    if isinstance(group, dict) and group.get('id') in driver_group_ids:
                        filtered_user_ids.add(user['id'])
                        break
            
            # Filter devices based on whether their driver is in the filtered users
            filtered_device_ids = {device_id for device_id, driver_id in device_to_driver.items() 
                                  if driver_id in filtered_user_ids}
            
            # Filter the devices list
            devices = [device for device in devices if device['id'] in filtered_device_ids]
        
        # Format devices for frontend
        formatted_devices = [{
            'id': device['id'],
            'name': device['name'],
            'serialNumber': device.get('serialNumber', ''),
        } for device in devices if device.get('name')]
        
        formatted_devices.sort(key=lambda x: x['name'].lower())
        
        return jsonify(formatted_devices)
    except Exception as e:
        print(f"Error in get_devices: {str(e)}")
        traceback.print_exc()  # Add import at the top if needed
        return jsonify({"error": str(e)}), 500

@app.route('/api/drivers', methods=['GET'])
def get_drivers():
    """Endpoint to fetch available drivers from Geotab, with optional driver group filtering."""
    try:
        driver_group_ids = request.args.getlist('driverGroupIds[]')

        # Get all device status info to find drivers and their associated devices
        device_status_info = get_geotab_data("DeviceStatusInfo")

        # Get all users in one call
        all_users = get_geotab_data("User")
        # Create a map of user IDs to user data for quick lookup
        user_map = {user['id']: user for user in all_users}

        # Create a dictionary to map driver IDs to their device IDs and user information
        driver_device_user_map = {}
        for status in device_status_info:
            if not status.get('driver') or not isinstance(status['driver'], dict) or not status['driver'].get('id'):
                continue
            driver_id = status['driver']['id']
            device_id = status['device']['id']
            if driver_id in user_map:
                driver_device_user_map[driver_id] = {
                    'user': user_map[driver_id],
                    'deviceId': device_id
                }

        # Filter drivers based on driver groups
        filtered_drivers_data = []
        for driver_id, data in driver_device_user_map.items():
            user = data['user']
            device_id = data['deviceId']
            if driver_group_ids:
                user_groups = user.get('driverGroups', [])
                if any(group.get('id') in driver_group_ids for group in user_groups):
                    filtered_drivers_data.append({'user': user, 'deviceId': device_id})
            else:
                filtered_drivers_data.append({'user': user, 'deviceId': device_id})

        # Format driver data with device ID
        drivers = [{
            'id': data['deviceId'],
            'name': f"{data['user'].get('firstName', '')} {data['user'].get('lastName', '')}".strip()
                    or data['user'].get('name', 'Unknown Driver'),
            'employeeNo': data['user'].get('employeeNo', '')
        } for data in filtered_drivers_data]

        # Sort by name
        drivers.sort(key=lambda x: x['name'].lower())

        print(f"Returning {len(drivers)} drivers")
        return jsonify(drivers)
    except Exception as e:
        print(f"Error in get_drivers: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/report-status')
def check_status():
    return jsonify(report_status)

@app.route('/api/download-report')
def download_report():
    if report_status["status"] == "complete" and os.path.exists(report_file_path):
        return send_file(report_file_path, as_attachment=True)
    else:
        return jsonify({"status": "error", "message": "Report not ready or not found"}), 404

if __name__ == '__main__':
    
    # Start the Flask app
    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)